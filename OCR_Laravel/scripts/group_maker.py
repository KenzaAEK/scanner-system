#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import random
import argparse
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT


# ───────────────────────────────────────────────────────── helpers ──

def shuffle_students(students):
    """Retourne une copie mélangée de la liste."""
    students = students.copy()
    random.shuffle(students)
    return students


def create_balanced_groups(students, group_size: int):
    """Retourne une liste de groupes équilibrés."""
    shuffled = shuffle_students(students)
    total    = len(shuffled)
    n_groups = (total + group_size - 1) // group_size  # arrondi sup

    groups = []
    for i in range(n_groups):
        start, end = i * group_size, min((i + 1) * group_size, total)
        members    = shuffled[start:end]
        if members:
            groups.append({'numero': i + 1, 'membres': members, 'taille': len(members)})

    # si dernier groupe trop petit → redistribution
    if len(groups) > 1 and groups[-1]['taille'] < group_size // 2:
        last = groups.pop()
        for idx, member in enumerate(last['membres']):
            target = groups[idx % len(groups)]
            target['membres'].append(member)
            target['taille'] += 1

    return groups


# ────────────────────────────────────────── Excel & PDF builders ──

def create_groups_excel(groups, output_path):
    """Crée un fichier Excel détaillant les groupes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Groupes TP"

    # styles
    border = Border(*(Side(style='thin') for _ in range(4)))
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", start_color="366092")

    ws.merge_cells('A1:C1')
    cell = ws['A1']
    cell.value, cell.font, cell.fill = "RÉPARTITION DES GROUPES DE TP", header_font, header_fill
    cell.alignment, cell.border = Alignment(horizontal='center'), border

    # infos générales
    infos = [
        ("Date de création:", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Nombre total d'étudiants:", sum(g['taille'] for g in groups)),
        ("Nombre de groupes:", len(groups)),
    ]
    start_row = 3
    for offset, (k, v) in enumerate(infos):
        ws[f'A{start_row+offset}'].value = k
        ws[f'A{start_row+offset}'].font  = Font(bold=True)
        ws[f'B{start_row+offset}'].value = v

    row = start_row + len(infos) + 2
    roles = ("Chef de groupe", "Secrétaire", "Membre")

    for grp in groups:
        # entête groupe
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        gcell = ws[f'A{row}']
        gcell.value = f"GROUPE {grp['numero']} ({grp['taille']} membres)"
        gcell.font  = Font(bold=True)
        gcell.fill  = PatternFill("solid", start_color="D9E1F2")
        gcell.alignment = Alignment(horizontal='center')
        gcell.border = border
        row += 1

        # sous-entêtes
        for col, txt in zip(("A", "B", "C"), ("N°", "Nom et Prénom", "Rôle")):
            c = ws[f'{col}{row}']
            c.value, c.font, c.border = txt, Font(bold=True), border
            c.fill = PatternFill("solid", start_color="F2F2F2")
            c.alignment = Alignment(horizontal='center')
        row += 1

        # membres
        for idx, name in enumerate(grp['membres'], 1):
            ws[f'A{row}'].value = idx
            ws[f'B{row}'].value = name
            ws[f'C{row}'].value = roles[min(idx - 1, 2)]
            for col in ("A", "B", "C"):
                ws[f'{col}{row}'].border = border
                ws[f'{col}{row}'].alignment = Alignment(horizontal='center' if col != 'B' else 'left')
            row += 1
        row += 1

    # largeur colonnes
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print("[OK] Excel créé :", output_path)
    return output_path


def create_groups_pdf(groups, output_path):
    """Crée un PDF récapitulatif."""
    doc = SimpleDocTemplate(output_path, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    title = Paragraph("RÉPARTITION DES GROUPES DE TP",
                      ParagraphStyle('t', parent=styles['Heading1'],
                                     alignment=TA_CENTER, textColor=colors.navy))

    info = [
        ['Date :', datetime.now().strftime("%d/%m/%Y %H:%M")],
        ["Total étudiants :", str(sum(g['taille'] for g in groups))],
        ["Nombre de groupes :", str(len(groups))],
    ]
    info_tbl = Table(info, colWidths=[5*cm, 4*cm])
    info_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))

    content = [title, Spacer(1, 12), info_tbl, Spacer(1, 24)]
    roles = ("Chef de groupe", "Secrétaire", "Membre")

    for grp in groups:
        gtitle = Paragraph(f"GROUPE {grp['numero']} ({grp['taille']} membres)",
                           ParagraphStyle('gt', parent=styles['Heading2'],
                                          textColor=colors.darkblue, spaceBefore=10))
        content.append(gtitle)

        data = [['N°', 'Nom et Prénom', 'Rôle']]
        for idx, name in enumerate(grp['membres'], 1):
            data.append([idx, name, roles[min(idx - 1, 2)]])

        table = Table(data, colWidths=[1.5*cm, 8*cm, 3*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.navy),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        content.extend([table, Spacer(1, 12)])

    doc.build(content)
    print("[OK] PDF créé  :", output_path)
    return output_path


# ────────────────────────────────────────────────────────── CLI ──

def main():
    parser = argparse.ArgumentParser(description="Génère des groupes d'étudiants.")
    parser.add_argument('input_file', help='Fichier texte (un étudiant par ligne)')
    parser.add_argument('group_size', type=int, help='Taille du groupe (2-8)')
    parser.add_argument('excel_output', help='Chemin du fichier Excel')
    parser.add_argument('pdf_output', help='Chemin du fichier PDF')
    args = parser.parse_args()

    with open(args.input_file, encoding='utf-8') as f:
        students = [s.strip() for s in f if s.strip()]

    if not (2 <= args.group_size <= 8):
        raise SystemExit("La taille du groupe doit être entre 2 et 8.")

    groups = create_balanced_groups(students, args.group_size)
    if not groups:
        raise SystemExit("Impossible de créer les groupes.")

    create_groups_excel(groups, args.excel_output)
    create_groups_pdf(groups, args.pdf_output)
    print("[SUCCESS] Groupes générés avec succès.")


if __name__ == "__main__":
    main()
