import os
import random
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from datetime import datetime

def shuffle_students(students_list):
    """Mélange aléatoire de la liste des étudiants"""
    shuffled = students_list.copy()
    random.shuffle(shuffled)
    return shuffled

def create_balanced_groups(students_list, group_size):
    """Création de groupes équilibrés"""
    try:
        # Mélange des étudiants
        shuffled_students = shuffle_students(students_list)
        
        # Calcul du nombre de groupes
        total_students = len(shuffled_students)
        num_groups = (total_students + group_size - 1) // group_size  # Division avec arrondi supérieur
        
        groups = []
        
        # Distribution équilibrée
        for i in range(num_groups):
            group = []
            start_idx = i * group_size
            end_idx = min(start_idx + group_size, total_students)
            
            for j in range(start_idx, end_idx):
                if j < len(shuffled_students):
                    group.append(shuffled_students[j])
            
            if group:  # Ajouter seulement si le groupe n'est pas vide
                groups.append({
                    'numero': i + 1,
                    'membres': group,
                    'taille': len(group)
                })
        
        # Redistribution si le dernier groupe est trop petit
        if len(groups) > 1 and groups[-1]['taille'] < group_size // 2:
            # Redistribuer les membres du dernier groupe
            last_group = groups.pop()
            for i, member in enumerate(last_group['membres']):
                target_group = groups[i % len(groups)]
                target_group['membres'].append(member)
                target_group['taille'] += 1
        
        return groups
        
    except Exception as e:
        print(f"Erreur création groupes: {e}")
        return []

def create_groups_excel(groups, output_path):
    """Création du fichier Excel avec les groupes"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Groupes TP"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        group_header_font = Font(bold=True, size=11)
        group_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # En-tête principal
        ws.merge_cells('A1:C1')
        header_cell = ws['A1']
        header_cell.value = "RÉPARTITION DES GROUPES DE TP"
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.border = border
        
        # Informations générales
        current_row = 3
        ws[f'A{current_row}'] = "Date de création:"
        ws[f'B{current_row}'] = datetime.now().strftime("%d/%m/%Y %H:%M")
        ws[f'A{current_row + 1}'] = "Nombre total d'étudiants:"
        ws[f'B{current_row + 1}'] = sum(group['taille'] for group in groups)
        ws[f'A{current_row + 2}'] = "Nombre de groupes:"
        ws[f'B{current_row + 2}'] = len(groups)
        
        # Style des informations
        for i in range(3):
            ws[f'A{current_row + i}'].font = Font(bold=True)
        
        current_row += 5
        
        # Groupes
        for group in groups:
            # En-tête du groupe
            ws.merge_cells(f'A{current_row}:C{current_row}')
            group_header = ws[f'A{current_row}']
            group_header.value = f"GROUPE {group['numero']} ({group['taille']} membres)"
            group_header.font = group_header_font
            group_header.fill = group_fill
            group_header.alignment = Alignment(horizontal='center', vertical='center')
            group_header.border = border
            
            current_row += 1
            
            # En-têtes colonnes
            ws[f'A{current_row}'] = "N°"
            ws[f'B{current_row}'] = "Nom et Prénom"
            ws[f'C{current_row}'] = "Rôle"
            
            for col in ['A', 'B', 'C']:
                cell = ws[f'{col}{current_row}']
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 1
            
            # Membres du groupe
            roles = ["Chef de groupe", "Secrétaire", "Membre"]
            for i, member in enumerate(group['membres']):
                ws[f'A{current_row}'] = i + 1
                ws[f'B{current_row}'] = member
                ws[f'C{current_row}'] = roles[min(i, len(roles) - 1)]
                
                for col in ['A', 'B', 'C']:
                    cell = ws[f'{col}{current_row}']
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center' if col != 'B' else 'left', vertical='center')
                
                current_row += 1
            
            current_row += 1  # Espace entre groupes
        
        # Ajustement des largeurs
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        
        # Création d'une feuille de statistiques
        stats_ws = wb.create_sheet("Statistiques")
        
        stats_data = [
            ["STATISTIQUES DES GROUPES"],
            [""],
            ["Répartition par taille de groupe:"]
        ]
        
        # Calcul des statistiques
        size_distribution = {}
        for group in groups:
            size = group['taille']
            size_distribution[size] = size_distribution.get(size, 0) + 1
        
        for size, count in sorted(size_distribution.items()):
            stats_data.append([f"Groupes de {size} membres:", count])
        
        stats_data.extend([
            [""],
            ["Taille moyenne:", round(sum(group['taille'] for group in groups) / len(groups), 1) if groups else 0],
            ["Taille minimale:", min(group['taille'] for group in groups) if groups else 0],
            ["Taille maximale:", max(group['taille'] for group in groups) if groups else 0],
        ])
        
        # Insertion des statistiques
        for row_idx, row_data in enumerate(stats_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = stats_ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
                elif len(row_data) > 1 and col_idx == 1:
                    cell.font = Font(bold=True)
        
        # Ajustement largeurs statistiques
        stats_ws.column_dimensions['A'].width = 25
        stats_ws.column_dimensions['B'].width = 10
        
        # Sauvegarde
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        
        print(f"✅ Excel groupes créé: {output_path}")
        return output_path
        
    except Exception as e:
        print(f"❌ Erreur création Excel groupes: {e}")
        return None

def create_groups_pdf(groups, output_path):
    """Création du PDF avec les groupes"""
    try:
        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        styles = getSampleStyleSheet()
        
        # Styles personnalisés
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.navy
        )
        
        group_title_style = ParagraphStyle(
            'GroupTitle',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=10,
            spaceBefore=15,
            alignment=TA_LEFT,
            textColor=colors.darkblue
        )
        
        story = []
        
        # Titre principal
        story.append(Paragraph("RÉPARTITION DES GROUPES DE TP", title_style))
        story.append(Spacer(1, 20))
        
        # Informations générales
        info_data = [
            ['Date de création:', datetime.now().strftime("%d/%m/%Y %H:%M")],
            ['Nombre total d\'étudiants:', str(sum(group['taille'] for group in groups))],
            ['Nombre de groupes:', str(len(groups))],
            ['Taille moyenne par groupe:', str(round(sum(group['taille'] for group in groups) / len(groups), 1)) if groups else '0']
        ]
        
        info_table = Table(info_data, colWidths=[5*cm, 4*cm])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(info_table)
        story.append(Spacer(1, 30))
        
        # Groupes
        for group in groups:
            # Titre du groupe
            story.append(Paragraph(f"GROUPE {group['numero']} ({group['taille']} membres)", group_title_style))
            
            # Tableau des membres
            group_data = [['N°', 'Nom et Prénom', 'Rôle']]
            
            roles = ["Chef de groupe", "Secrétaire", "Membre"]
            for i, member in enumerate(group['membres']):
                group_data.append([
                    str(i + 1),
                    member,
                    roles[min(i, len(roles) - 1)]
                ])
            
            group_table = Table(group_data, colWidths=[1.5*cm, 8*cm, 3*cm])
            group_table.setStyle(TableStyle([
                # En-tête
                ('BACKGROUND', (0, 0), (-1, 0), colors.navy),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                
                # Contenu
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                
                # Bordures et alignement
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # N°
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),    # Noms
                ('ALIGN', (2, 0), (2, -1), 'CENTER'),  # Rôles
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                
                # Alternance de couleurs
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            story.append(group_table)
            story.append(Spacer(1, 20))
        
        # Instructions
        story.append(Spacer(1, 20))
        story.append(Paragraph("<b>Instructions:</b>", styles['Heading3']))
        
        instructions = [
            "• Le chef de groupe coordonne les activités et représente le groupe",
            "• Le secrétaire prend notes et gère la communication",
            "• Tous les membres participent activement aux travaux",
            "• Les rôles peuvent être alternés selon les séances"
        ]
        
        for instruction in instructions:
            story.append(Paragraph(instruction, styles['Normal']))
        
        # Construction du PDF
        doc.build(story)
        
        print(f"✅ PDF groupes créé: {output_path}")
        return output_path
        
    except Exception as e:
        print(f"❌ Erreur création PDF groupes: {e}")
        return None

def create_student_groups(students_list, group_size, excel_output, pdf_output):
    """Fonction principale de création des groupes"""
    try:
        # Validation des paramètres
        if not students_list:
            raise ValueError("La liste des étudiants est vide")
        
        if group_size < 2 or group_size > 8:
            raise ValueError("La taille des groupes doit être entre 2 et 8")
        
        # Nettoyage de la liste
        clean_students = [name.strip() for name in students_list if name.strip()]
        
        if len(clean_students) < 2:
            raise ValueError("Il faut au moins 2 étudiants pour former des groupes")
        
        print(f"🔄 Création de groupes pour {len(clean_students)} étudiants (taille: {group_size})")
        
        # Création des groupes
        groups = create_balanced_groups(clean_students, group_size)
        
        if not groups:
            raise ValueError("Erreur lors de la création des groupes")
        
        print(f"👥 {len(groups)} groupes créés")
        
        # Génération des fichiers
        excel_result = create_groups_excel(groups, excel_output)
        pdf_result = create_groups_pdf(groups, pdf_output)
        
        return excel_result, pdf_result
        
    except Exception as e:
        print(f"❌ Erreur création groupes: {e}")
        return None, None

# Test du module
if __name__ == "__main__":
    test_students = [
        "Ahmed ALAMI", "Fatima BENALI", "Youssef TAZI", "Khadija FASSI",
        "Omar IDRISSI", "Zineb BERRADA", "Hassan CHERKAOUI", "Salma BENNANI",
        "Karim HAKIMI", "Nadia LAHLOU", "Rachid SEFRIOUI", "Laila SABRI"
    ]
    
    excel_test = "test_groupes.xlsx"
    pdf_test = "test_groupes.pdf"
    
    excel_result, pdf_result = create_student_groups(test_students, 3, excel_test, pdf_test)
    
    if excel_result and pdf_result:
        print("Test réussi!")
    else:
        print("Test échoué!")

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description='Créer des groupes d\'étudiants')
    parser.add_argument('input_file', help='Fichier texte avec étudiants (un par ligne)')
    parser.add_argument('group_size', type=int, help='Taille des groupes')
    parser.add_argument('excel_output', help='Chemin de sortie Excel')
    parser.add_argument('pdf_output', help='Chemin de sortie PDF')
    
    args = parser.parse_args()
    
    # Lire les étudiants du fichier
    with open(args.input_file, 'r') as f:
        students = [line.strip() for line in f.readlines() if line.strip()]
    
    excel_result, pdf_result = create_student_groups(
        students,
        args.group_size,
        args.excel_output,
        args.pdf_output
    )
    
    if excel_result and pdf_result:
        print(excel_result)
        print(pdf_result)
    else:
        exit(1)