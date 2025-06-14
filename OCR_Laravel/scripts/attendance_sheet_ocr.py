import cv2
import numpy as np
import pytesseract
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from fpdf import FPDF
import os
import tempfile
from PIL import Image

def extract_students_from_image_ocr(image_path):
    """
    Extrait la liste des étudiants d'une image de tableau avec OCR
    """
    # Vérification de l'image
    img = cv2.imread(image_path)
    if img is None:
        raise FileNotFoundError(f"Image introuvable : {image_path}")

    # Prétraitement
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)

    # Détection des lignes verticales et horizontales
    vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, img.shape[0] // 30))
    vertical_lines = cv2.morphologyEx(binary, cv2.MORPH_OPEN, vertical_kernel, iterations=2)

    horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (img.shape[1] // 30, 1))
    horizontal_lines = cv2.morphologyEx(binary, cv2.MORPH_OPEN, horizontal_kernel, iterations=2)

    table_mask = cv2.add(vertical_lines, horizontal_lines)

    # Détection des cellules par contours
    contours, _ = cv2.findContours(table_mask, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    cells = [cv2.boundingRect(c) for c in contours if cv2.contourArea(c) > 500]
    cells = sorted(cells, key=lambda b: (b[1], b[0]))  # Trier par ligne, puis colonne

    # Regrouper les cellules par lignes
    rows = []
    current_row = []
    last_y = -100
    tolerance = 25

    for (x, y, w, h) in cells:
        if abs(y - last_y) > tolerance:
            if current_row:
                rows.append(current_row)
            current_row = [(x, y, w, h)]
            last_y = y
        else:
            current_row.append((x, y, w, h))
    if current_row:
        rows.append(current_row)

    # Extraire le texte de chaque cellule
    extracted_data = []
    for i, row in enumerate(rows):
        row = sorted(row, key=lambda b: b[0])  # Tri horizontal
        row_data = []
        for j, (x, y, w, h) in enumerate(row):
            # Extraire image cellule
            cell_img = img[y:y+h, x:x+w]
            cell_gray = cv2.cvtColor(cell_img, cv2.COLOR_BGR2GRAY)
            cell_thresh = cv2.threshold(cell_gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]

            # OCR pour la cellule
            config = "--psm 7"
            text = pytesseract.image_to_string(cell_thresh, config=config).strip()
            row_data.append(text)
        
        if row_data:  # Éviter les lignes vides
            extracted_data.append(row_data)

    return extracted_data

def generate_attendance_pdf_from_ocr(image_path, class_name, num_sessions, output_path, include_notes=False, custom_header=""):
    """
    Génère une feuille de présence PDF à partir d'une image de liste d'étudiants
    """
    try:
        # Extraire les données de l'image
        extracted_data = extract_students_from_image_ocr(image_path)
        
        if not extracted_data:
            raise ValueError("Aucune donnée extraite de l'image")

        # Traiter les données extraites pour identifier les étudiants
        students = []
        for row in extracted_data:
            if len(row) >= 2:  # Au moins 2 colonnes (nom, prénom ou code, nom)
                # Nettoyer et formater les données
                cleaned_row = [cell.strip() for cell in row if cell.strip()]
                if cleaned_row:
                    # Joindre les données de la ligne comme nom complet
                    student_name = " ".join(cleaned_row[:3])  # Prendre les 3 premières colonnes max
                    students.append(student_name)

        if not students:
            raise ValueError("Aucun étudiant identifié dans l'image")

        # Générer le PDF
        pdf = FPDF(orientation='L', unit='mm', format='A4')
        pdf.add_page()
        
        # En-tête personnalisé
        if custom_header:
            pdf.set_font("Arial", "B", 12)
            pdf.cell(0, 8, custom_header, ln=True, align="C")
            pdf.ln(3)
        
        # Titre principal
        pdf.set_font("Arial", "B", 14)
        pdf.cell(0, 10, f"Feuille de présence - {class_name}", ln=True, align="C")
        pdf.ln(5)

        # Calculer les largeurs des colonnes
        base_width = 50  # Largeur pour le nom
        session_width = max(15, (297 - base_width - 20) // num_sessions)  # Répartir l'espace restant
        
        # En-têtes
        headers = ["Nom Complet"] + [f"S{i+1}" for i in range(num_sessions)]
        if include_notes:
            headers.append("Notes")
        
        col_widths = [base_width] + [session_width] * num_sessions
        if include_notes:
            col_widths.append(40)
        
        row_height = 6

        # Dessiner les en-têtes
        pdf.set_font("Arial", "B", 9)
        for header, width in zip(headers, col_widths):
            pdf.cell(width, row_height, header, border=1, align="C")
        pdf.ln()

        # Dessiner les lignes d'étudiants
        pdf.set_font("Arial", "", 8)
        for i, student in enumerate(students):
            # Nom de l'étudiant
            pdf.cell(col_widths[0], row_height, student[:30], border=1)  # Limiter à 30 chars
            
            # Cases de présence
            for j in range(num_sessions):
                pdf.cell(col_widths[j+1], row_height, "", border=1)
            
            # Colonne notes si demandée
            if include_notes:
                pdf.cell(col_widths[-1], row_height, "", border=1)
            
            pdf.ln()

        # Informations supplémentaires
        pdf.ln(5)
        pdf.set_font("Arial", "", 8)
        pdf.cell(0, 4, f"Nombre d'étudiants : {len(students)}", ln=True)
        pdf.cell(0, 4, f"Nombre de séances : {num_sessions}", ln=True)

        # Sauvegarder le PDF
        pdf.output(output_path)
        return output_path

    except Exception as e:
        raise Exception(f"Erreur lors de la génération de la feuille de présence : {str(e)}")

def generate_attendance_excel_from_ocr(image_path, class_name, num_sessions, output_path, include_notes=False):
    """
    Génère une feuille de présence Excel à partir d'une image de liste d'étudiants
    """
    try:
        # Extraire les données de l'image
        extracted_data = extract_students_from_image_ocr(image_path)
        
        if not extracted_data:
            raise ValueError("Aucune donnée extraite de l'image")

        # Traiter les données extraites
        students = []
        for row in extracted_data:
            if len(row) >= 1:
                cleaned_row = [cell.strip() for cell in row if cell.strip()]
                if cleaned_row:
                    student_name = " ".join(cleaned_row[:3])  # Prendre les 3 premières colonnes max
                    students.append(student_name)

        if not students:
            raise ValueError("Aucun étudiant identifié dans l'image")

        # Créer le fichier Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Présence_{class_name}"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))

        # En-têtes
        headers = ["Nom Complet"] + [f"Séance {i+1}" for i in range(num_sessions)]
        if include_notes:
            headers.append("Notes")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 30  # Nom
        for i in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 12

        # Remplir avec les étudiants
        for row, student in enumerate(students, 2):
            cell = ws.cell(row=row, column=1, value=student)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Cases vides pour les séances
            for col in range(2, len(headers) + 1):
                cell = ws.cell(row=row, column=col, value="")
                cell.border = border

        # Sauvegarder
        wb.save(output_path)
        return output_path

    except Exception as e:
        raise Exception(f"Erreur lors de la génération du fichier Excel : {str(e)}")

# Fonction compatible avec votre interface Streamlit existante
def generate_attendance_from_image(image_path, class_name, num_sessions, output_format="pdf", 
                                 include_notes=False, custom_header=""):
    """
    Fonction principale compatible avec votre interface Streamlit
    """
    output_dir = "generated_files"
    os.makedirs(output_dir, exist_ok=True)
    
    if output_format.lower() == "pdf":
        output_path = os.path.join(output_dir, f"presence_{class_name}_{num_sessions}seances.pdf")
        return generate_attendance_pdf_from_ocr(image_path, class_name, num_sessions, 
                                              output_path, include_notes, custom_header)
    else:
        output_path = os.path.join(output_dir, f"presence_{class_name}_{num_sessions}seances.xlsx")
        return generate_attendance_excel_from_ocr(image_path, class_name, num_sessions, 
                                                output_path, include_notes)

if __name__ == "__main__":
    # Test
    image_path = "./images/liste11.jpg"
    class_name = "GI-S5"
    num_sessions = 6
    
    try:
        pdf_result = generate_attendance_from_image(image_path, class_name, num_sessions, "pdf")
        print(f"✅ PDF généré : {pdf_result}")
        
        excel_result = generate_attendance_from_image(image_path, class_name, num_sessions, "excel")
        print(f"✅ Excel généré : {excel_result}")
    except Exception as e:
        print(f"❌ Erreur : {e}")