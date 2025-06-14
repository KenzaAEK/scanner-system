import os
import cv2
import pandas as pd
from paddleocr import PaddleOCR
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Initialiser PaddleOCR en français
ocr = PaddleOCR(use_angle_cls=True, lang='fr')

def group_by_y(entries, tolerance=15):
    entries = sorted(entries, key=lambda e: e['y'])
    groups = []
    current = []
    last_y = None
    for entry in entries:
        if last_y is None or abs(entry['y'] - last_y) < tolerance:
            current.append(entry)
        else:
            groups.append(current)
            current = [entry]
        last_y = entry['y']
    if current:
        groups.append(current)
    return groups

def extract_table(image_path: str, min_confidence=0.5):
    if not os.path.exists(image_path):
        raise FileNotFoundError(f"Image introuvable : {image_path}")

    import pprint
    pprint.pprint(result[0][:5])
    result = ocr.ocr(image_path)

    entries = []

    for line in result[0]:
        try:
            box = line[0]
            content = line[1]

            # ✅ Cas standard: tuple (text, confidence)
            if isinstance(content, (list, tuple)) and len(content) == 2:
                text, confidence = content
            # ✅ Cas fallback : liste à 1 seul élément ou tuple cassé
            elif isinstance(content, (list, tuple)) and len(content) == 1:
                text = content[0] if content[0] else ""
                confidence = 1.0
            # ✅ Cas dégradé : juste une string
            elif isinstance(content, str):
                text = content
                confidence = 1.0
            # 🚫 Cas non reconnu
            else:
                raise ValueError("Format OCR inattendu")

            # Ignore les textes vides ou pas assez fiables
            if not text.strip() or confidence < min_confidence:
                continue

            # Calculer la position
            x = min(pt[0] for pt in box)
            y = min(pt[1] for pt in box)
            entries.append({'text': text.strip(), 'x': x, 'y': y})

        except Exception as e:
            print(f"⚠️ Ligne ignorée : {e}")





    lines = group_by_y(entries)
    table = []
    for line in lines:
        sorted_line = sorted(line, key=lambda e: e['x'])
        row = [cell['text'] for cell in sorted_line]
        table.append(row)

    if not table:
        raise ValueError("⚠️ Aucun contenu OCR détecté dans l'image.")

    max_cols = max(len(row) for row in table)
    table = [row + [""] * (max_cols - len(row)) for row in table]
    return table

def save_table_to_excel(table, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tableau OCR"

    # Styles
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for i, row in enumerate(table):
        for j, val in enumerate(row):
            cell = ws.cell(row=i+1, column=j+1, value=val)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            if i == 0:
                cell.font = header_font
                cell.fill = header_fill

    # Ajustement des colonnes
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path

def image_to_excel_converter_local(image_path: str, output_path: str,
                                   min_confidence: float = 0.5) -> str:
    print(f"🔍 Lecture de l'image : {image_path}")
    table = extract_table(image_path, min_confidence)
    if not table or all(len(row) <= 1 for row in table):
        raise ValueError("Aucun tableau structuré détecté dans l'image.")
    
    final_path = save_table_to_excel(table, output_path)
    print(f"✅ Export Excel terminé : {final_path}")
    return final_path

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description='Extraire un tableau vers Excel')
    parser.add_argument('image_path', help='Chemin de l\'image')
    parser.add_argument('output_path', help='Chemin de sortie Excel')
    parser.add_argument('--min_confidence', type=float, default=0.5, help='Confiance minimale OCR')
    
    args = parser.parse_args()
    
    result = image_to_excel_converter_local(
        args.image_path,
        args.output_path,
        args.min_confidence
    )
    
    print(result)
