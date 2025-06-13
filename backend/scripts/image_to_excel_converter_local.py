import os
import easyocr
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

reader = easyocr.Reader(['fr'], gpu=False)

def group_by_y(entries, tolerance=15):
    entries = sorted(entries, key=lambda e: e['y'])
    grouped = []
    current = []
    last_y = None
    for e in entries:
        if last_y is None or abs(e['y'] - last_y) < tolerance:
            current.append(e)
        else:
            grouped.append(current)
            current = [e]
        last_y = e['y']
    if current:
        grouped.append(current)
    return grouped

def extract_easyocr_table(image_path):
    results = reader.readtext(image_path)
    entries = []
    for box, text, confidence in results:
        if not text.strip():
            continue
        x = min(p[0] for p in box)
        y = min(p[1] for p in box)
        entries.append({'text': text.strip(), 'x': x, 'y': y})
    return entries

def save_table_to_excel(table, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "OCR_Table"

    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for i, row in enumerate(table):
        for j, val in enumerate(row):
            cell = ws.cell(row=i+1, column=j+1, value=val)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            if i == 0:
                cell.font = header_font
                cell.fill = header_fill

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    wb.save(output_path)
    return output_path

def image_to_excel_converter_local(image_path, output_path):
    print(f"📷 Lecture de l'image avec EasyOCR : {image_path}")
    entries = extract_easyocr_table(image_path)

    if not entries:
        raise ValueError("❌ Aucun texte détecté.")

    lines = group_by_y(entries)
    table = []
    for line in lines:
        sorted_line = sorted(line, key=lambda e: e['x'])
        row = [cell['text'] for cell in sorted_line]
        table.append(row)

    max_cols = max(len(row) for row in table)
    table = [row + [""] * (max_cols - len(row)) for row in table]

    return save_table_to_excel(table, output_path)

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description='Convertir image vers Excel avec OCR')
    parser.add_argument('image_path', help='Chemin de l\'image')
    parser.add_argument('output_path', help='Chemin de sortie Excel')
    
    args = parser.parse_args()
    
    result = image_to_excel_converter_local(args.image_path, args.output_path)
    print(result)
